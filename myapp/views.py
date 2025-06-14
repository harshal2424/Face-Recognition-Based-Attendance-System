import cv2
import os
import shutil
from django.conf import settings
from django.shortcuts import render, redirect, get_object_or_404
import numpy as np
from datetime import date, datetime
from .models import Students, Attendance, Course, Subject
from .forms import CourseForm
from .forms import StudentForm # Still needed by manage_course_students_view
# from django.forms import modelformset_factory # No longer needed
from django.utils.timezone import now
from django.http import HttpResponse, JsonResponse
from django.db import IntegrityError, transaction
from django.core.files.storage import FileSystemStorage
import face_recognition # USING THIS FOR RECOGNITION
from django.contrib import messages
from django.views.decorators.http import require_POST
from PIL import Image
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
# from ultralytics import YOLO # Removed
from deepface import DeepFace # USING THIS FOR DETECTION


def mark_attendance(request):
    if request.method == "POST":
        date_str = request.POST.get("date")
        course_id = request.POST.get("course")
        course = get_object_or_404(Course, id=course_id)
        students = Students.objects.all()
        for student in students:
            status = request.POST.get(f"status_{student.id}")
            photo = request.FILES.get(f"photo_{student.id}")
            if status:
                Attendance.objects.update_or_create(
                    date=date_str,
                    course=course,
                    student=student,
                    defaults={"status": status, "photo": photo},
                )
        return render(request, 'attendance_success.html', {"course": course, "date": date_str})

    courses = Course.objects.all()
    students = Students.objects.all()
    context = {
        "courses": courses,
        "students": students,
        "today": date.today().strftime('%Y-%m-%d'),
    }
    return render(request, 'mark_attendance.html', context)

def result(request):
    print("Executing result view (GET)" if request.method == "GET" else "Executing result view (POST)")
    if request.method == "GET":
        students = Students.objects.all()
        courses = Course.objects.all()
        context = {
            "students": students,
            "courses": courses,
            "today": date.today().strftime('%Y-%m-%d'),
            "MEDIA_URL": settings.MEDIA_URL,
        }
        return render(request, 'result.html', context)

    elif request.method == "POST":
        print("Handling POST request in result view")
        course_name = request.POST.get("course")
        print("Selected course:", course_name)
        try:
            course = Course.objects.get(name=course_name)
        except Course.DoesNotExist:
            return render(request, "attendance_failed.html", {"error": "Course not found"})

        attendance_date = request.POST.get("date")
        i = 1
        while True:
            student_name = request.POST.get(f"student_{i}")
            face_image_url = request.POST.get(f"face_image_url_{i}")
            if student_name is None or face_image_url is None: break
            if student_name == "Other": i += 1; continue
            try:
                student = Students.objects.get(name=student_name)
            except Students.DoesNotExist: i += 1; continue
            try:
                 relative_photo_path = face_image_url.replace(settings.MEDIA_URL, '', 1) if face_image_url.startswith(settings.MEDIA_URL) else face_image_url
                 Attendance.objects.update_or_create(
                     date=attendance_date, course=course, student=student,
                     defaults={"status": "Present", "photo": relative_photo_path}
                 )
            except IntegrityError: print(f"Attendance already exists for {student_name} on {attendance_date}")
            except Exception as e: print(f"Error saving attendance for {student_name}: {e}")
            i += 1
        messages.success(request, f"Attendance submitted from result page for {course.name} on {attendance_date}.")
        return redirect('manage_courses')

def attendance_success(request):
    return render(request, 'attendance_success.html')

def attendance_failed(request):
    context = {'error': request.GET.get('error', 'An unknown error occurred.')}
    return render(request, 'attendance_failed.html', context)

def login_view(request):
    if request.method == 'POST':
        login_type = request.POST.get('login_type')
        if login_type == 'admin':
             print("Admin login attempted (redirecting)")
             return redirect('admin_dashboard')
        elif login_type == 'instructor':
             print("Instructor login attempted (redirecting)")
             return redirect('instructor_dashboard')
        else:
             return render(request, 'login.html', {'error': 'Invalid login type or credentials'})
    else:
        return render(request, 'login.html', {})

def admin_dashboard(request):
    return render(request, 'admin_dashboard.html')


def create_course_view(request):
    if request.method == 'POST':
        course_form = CourseForm(request.POST)
        if course_form.is_valid():
            try:
                new_course = course_form.save()
                messages.success(request, f'Course "{new_course.name}" created successfully! You can now add students via Manage Courses.')
                print(f"Course created: {new_course.name} with ID: {new_course.id}")
                try:
                    course_db_dir = os.path.join(settings.MEDIA_ROOT, 'student_photo_db', str(new_course.id))
                    os.makedirs(course_db_dir, exist_ok=True)
                    print(f"Ensured DeepFace directory exists: {course_db_dir}")
                    messages.info(request, f"Student photo DB directory created for {new_course.name}. Add student photos (named STUDENT_ID.jpg) there.")
                except OSError as e:
                    messages.error(request, f"Could not create student photo directory for the course: {e}")
                    print(f"Error creating directory {course_db_dir}: {e}")
                return redirect('create_course')
            except Exception as e:
                 messages.error(request, f"An unexpected error occurred during saving: {e}")
                 print(f"Error in create_course_view saving course: {e}")
        else:
            print("Course Form Errors:", course_form.errors)
            messages.error(request, "Please correct the errors in the course details.")

    else: # GET request
        course_form = CourseForm()
    context = { 'course_form': course_form }
    return render(request, 'create_course.html', context)

def add_subject_view(request):
    courses = Course.objects.all()
    if request.method == 'POST':
        try:
            course_id = request.POST.get('selected_course')
            subject_name = request.POST.get('subject_name')
            teacher_name = request.POST.get('subject_teacher')

            if not all([course_id, subject_name, teacher_name]):
                messages.error(request, "All fields (Course, Subject Name, Teacher Name) are required.")
            else:
                try:
                    selected_course = Course.objects.get(id=course_id)
                    new_subject = Subject.objects.create(
                        name=subject_name,
                        teacher_name=teacher_name,
                        course=selected_course
                    )
                    messages.success(request, f'Successfully added subject "{subject_name}" to course "{selected_course.name}".')
                    return redirect('add_subject')
                except Course.DoesNotExist:
                    messages.error(request, "Selected course not found.")
                except IntegrityError:
                    try:
                        selected_course_name = Course.objects.get(id=course_id).name
                        messages.error(request, f'Subject "{subject_name}" already exists for the course "{selected_course_name}".')
                    except Course.DoesNotExist:
                         messages.error(request, f'Subject "{subject_name}" could not be added as the selected course was not found.')
                except Exception as e:
                    messages.error(request, f"An unexpected error occurred: {e}")
        except Exception as e:
            messages.error(request, f"An error occurred processing the form: {e}")

        context = {'courses': courses}
        return render(request, 'add_subject.html', context)
    else:
        context = {'courses': courses}
        return render(request, 'add_subject.html', context)


def instructor_dashboard_view(request):
    courses = Course.objects.all()
    context = {
        'courses': courses,
        'today': date.today().strftime('%Y-%m-%d'),
        'processed_data': None,
        'manual_correction_data': None,
        'processed_course_id': None
    }

    # --- Handle POST request for Manual Corrections ---
    if request.method == 'POST' and 'submit_manual_corrections' in request.POST:
        # ... (Keep existing manual correction logic as provided before) ...
        try:
            correction_date_str = request.POST.get('correction_date')
            correction_course_id = request.POST.get('correction_course_id')
            correction_date = datetime.strptime(correction_date_str, '%Y-%m-%d').date()
            course = get_object_or_404(Course, id=correction_course_id)
            corrections_made = 0
            unknowns_logged = 0

            with transaction.atomic():
                for key, value in request.POST.items():
                    if key.startswith('manual_assignment_'):
                        face_unique_id = key.replace('manual_assignment_', '')
                        if value and value != "UNKNOWN":
                            student_id = value
                            try:
                                student_to_mark = Students.objects.get(id=student_id)
                                obj, created = Attendance.objects.update_or_create(
                                    date=correction_date, course=course, student=student_to_mark,
                                    defaults={'status': 'Present'}
                                )
                                print(f"[Manual Correction] {'Created' if created else 'Updated'} Present record for {student_to_mark.name}")
                                corrections_made += 1
                            except Students.DoesNotExist: messages.error(request, f"Selected student ID {student_id} not found.")
                            except Exception as e: messages.error(request, f"Error applying correction for student ID {student_id}: {e}")
                        elif value == "UNKNOWN":
                            unknown_name_key = f'unknown_name_{face_unique_id}'
                            entered_name = request.POST.get(unknown_name_key, '').strip()
                            print(f"[Manual Correction] Face {face_unique_id} marked as UNKNOWN. Entered name: '{entered_name if entered_name else '(Not Provided)'}'")
                            unknowns_logged += 1

            if corrections_made > 0:
                messages.success(request, f"Applied {corrections_made} manual attendance correction(s) for {correction_date_str}.")
            if unknowns_logged > 0:
                 messages.info(request, f"Logged {unknowns_logged} face(s) marked as Unknown.")
            return redirect('instructor_dashboard')

        except Exception as e:
            messages.error(request, f"Error processing manual corrections: {e}")
            print(f"Error processing manual corrections: {e}")
            return redirect('instructor_dashboard')

    # --- Handle POST request for Initial Face Processing (Multiple Photos Version) ---
    elif request.method == 'POST':
        # --- ADDED DEBUG PRINTS ---
        print("\n--- instructor_dashboard POST Request Received ---")
        print("request.POST data:", request.POST)
        print("request.FILES data:", request.FILES)
        # --- END DEBUG PRINTS ---

        message = None
        error = None
        processed_data = None
        manual_correction_data = None
        processed_course_id = None
        all_extracted_faces_info = []
        processed_files_count = 0
        uploaded_files = [] # List to hold actual UploadedFile objects

        try:
            course_id = request.POST.get('selected_course')
            attendance_date_str = request.POST.get('attendance_date')
            # Get files from named inputs
            photo1 = request.FILES.get('group_photo_1')
            photo2 = request.FILES.get('group_photo_2')
            photo3 = request.FILES.get('group_photo_3')

            # Collect the files that were actually uploaded
            if photo1: uploaded_files.append(photo1)
            if photo2: uploaded_files.append(photo2)
            if photo3: uploaded_files.append(photo3)

            # --- ADDED DEBUG PRINTS ---
            print(f"[DEBUG] Course ID received: {course_id}")
            print(f"[DEBUG] Date Str received: {attendance_date_str}")
            print(f"[DEBUG] Number of files found in request.FILES: {len(uploaded_files)}")
            # --- END DEBUG PRINTS ---

            # Check if required fields are present AND at least one file was uploaded
            if not all([course_id, attendance_date_str]) or not uploaded_files:
                error = "Please select a course, date, and upload at least one group photo."
                print(f"[DEBUG] Validation FAILED: Course={course_id}, Date={attendance_date_str}, Files Found={bool(uploaded_files)}")
            else:
                print(f"[DEBUG] Validation PASSED. Proceeding...")
                selected_course = get_object_or_404(Course, id=course_id)
                processed_course_id = selected_course.id # Store for re-rendering dropdown
                attendance_date = datetime.strptime(attendance_date_str, '%Y-%m-%d').date()
                timestamp = datetime.now().strftime('%Y%m%d%H%M%S')

                # --- Load Known Student Encodings ONCE ---
                print("[DEBUG] Loading known student faces/encodings...")
                students_in_course = selected_course.students.all()
                known_face_encodings = []
                known_face_student_ids = []
                for student in students_in_course:
                    if student.photo and hasattr(student.photo, 'path') and student.photo.path and os.path.exists(student.photo.path):
                        try:
                            known_image = face_recognition.load_image_file(student.photo.path)
                            known_encodings = face_recognition.face_encodings(known_image)
                            if known_encodings:
                                known_face_encodings.append(known_encodings[0])
                                known_face_student_ids.append(student.id)
                            else: print(f"[WARNING] No face encoding for {student.name}")
                        except Exception as enc_err: print(f"[ERROR] Could not load/encode face for {student.name}: {enc_err}")
                    else: print(f"[WARNING] No photo path for {student.name}")

                if not known_face_encodings:
                     messages.warning(request,"Could not load any known student face encodings for comparison.")
                else:
                     print(f"[DEBUG] Loaded {len(known_face_encodings)} known encodings.")


                # --- Process EACH uploaded photo ---
                for photo_index, uploaded_file in enumerate(uploaded_files):
                    print(f"[DEBUG] Processing uploaded file {photo_index+1}: {uploaded_file.name}")
                    fs = FileSystemStorage(location=os.path.join(settings.MEDIA_ROOT, 'group_photos'))
                    filename = fs.save(f"{timestamp}_{photo_index}_{uploaded_file.name}", uploaded_file)
                    uploaded_file_path = fs.path(filename)
                    print(f"[DEBUG] Group photo saved to: {uploaded_file_path}")

                    # --- Detect & Extract Faces using DeepFace for THIS photo ---
                    try:
                        detector_backend = 'retinaface'
                        print(f"[DEBUG] Extracting faces from photo {photo_index+1} using DeepFace...")
                        extracted_data_single_photo = DeepFace.extract_faces(
                            img_path=uploaded_file_path, detector_backend=detector_backend,
                            enforce_detection=False, align=True
                        )
                        extracted_data_single_photo = [item for item in extracted_data_single_photo if 'face' in item and isinstance(item['face'], np.ndarray)]
                        print(f"[DEBUG] DeepFace extracted {len(extracted_data_single_photo)} faces from photo {photo_index+1}.")

                        for detection_index, item in enumerate(extracted_data_single_photo):
                            unique_face_id = f"{photo_index}_{detection_index}"
                            all_extracted_faces_info.append({
                                'unique_id': unique_face_id,
                                'face_array': item['face'],
                            })
                        processed_files_count += 1

                    except Exception as extract_err:
                        print(f"[ERROR] Failed to extract faces from {uploaded_file.name}: {extract_err}")
                        messages.error(request, f"Failed to extract faces from {uploaded_file.name}: {extract_err}")
                        continue

                # --- Recognize ALL extracted faces using face_recognition ---
                present_student_ids = set()
                unidentified_face_details = []
                identified_face_indices = set() # Track unique_ids identified

                if all_extracted_faces_info and known_face_encodings:
                    print(f"[DEBUG] Comparing {len(all_extracted_faces_info)} total detected faces using face_recognition...")
                    for face_info in all_extracted_faces_info:
                        unique_id = face_info['unique_id']
                        face_array_bgr_float = face_info['face_array']
                        match_found_for_this_face = False
                        try:
                            face_array_rgb_uint8 = cv2.cvtColor((face_array_bgr_float * 255).astype(np.uint8), cv2.COLOR_BGR2RGB)
                            unknown_encodings = face_recognition.face_encodings(face_array_rgb_uint8)

                            if unknown_encodings:
                                unknown_encoding = unknown_encodings[0]
                                matches = face_recognition.compare_faces(known_face_encodings, unknown_encoding, tolerance=0.6)
                                face_distances = face_recognition.face_distance(known_face_encodings, unknown_encoding)
                                best_match_index = -1
                                if True in matches:
                                    best_match_index = np.argmin(face_distances)

                                if best_match_index != -1:
                                    matched_student_id = known_face_student_ids[best_match_index]
                                    present_student_ids.add(matched_student_id)
                                    identified_face_indices.add(unique_id)
                                    match_found_for_this_face = True
                                    print(f"[DEBUG] face_recognition matched face {unique_id} to student ID {matched_student_id} (Distance: {face_distances[best_match_index]:.4f})")

                        except Exception as rec_err: print(f"[ERROR] Error processing face {unique_id} with face_recognition: {rec_err}")

                        # If no match, prepare for manual correction UI
                        if not match_found_for_this_face:
                             try:
                                 final_unid_folder = os.path.join(settings.MEDIA_ROOT, 'processed', 'unidentified_faces', attendance_date_str, str(selected_course.id))
                                 os.makedirs(final_unid_folder, exist_ok=True)
                                 final_filename = f"unid_face_{unique_id}.jpg"
                                 relative_path = os.path.join('processed', 'unidentified_faces', attendance_date_str, str(selected_course.id), final_filename).replace(os.sep, "/")
                                 final_save_path = os.path.join(settings.MEDIA_ROOT, relative_path)
                                 img_to_save = Image.fromarray(face_array_rgb_uint8)
                                 img_to_save.save(final_save_path)
                                 face_url = f'{settings.MEDIA_URL}{relative_path}'
                                 unidentified_face_details.append({'url': face_url, 'index': unique_id})
                             except Exception as save_e: print(f"Error saving unidentified face {unique_id}: {save_e}")

                elif not known_face_encodings and all_extracted_faces_info:
                     print("[WARNING] Known encodings missing, saving all detected faces as unidentified.")
                     final_unid_folder = os.path.join(settings.MEDIA_ROOT, 'processed', 'unidentified_faces', attendance_date_str, str(selected_course.id))
                     os.makedirs(final_unid_folder, exist_ok=True)
                     for face_info in all_extracted_faces_info:
                        # ... (Code to save all faces as unidentified, same as loop above) ...
                         unique_id = face_info['unique_id']
                         face_array_bgr_float = face_info['face_array']
                         face_array_rgb_uint8 = cv2.cvtColor((face_array_bgr_float * 255).astype(np.uint8), cv2.COLOR_BGR2RGB)
                         try:
                             final_filename = f"unid_face_{unique_id}.jpg"
                             relative_path = os.path.join('processed', 'unidentified_faces', attendance_date_str, str(selected_course.id), final_filename).replace(os.sep, "/")
                             final_save_path = os.path.join(settings.MEDIA_ROOT, relative_path)
                             img_to_save = Image.fromarray(face_array_rgb_uint8)
                             img_to_save.save(final_save_path)
                             face_url = f'{settings.MEDIA_URL}{relative_path}'
                             unidentified_face_details.append({'url': face_url, 'index': unique_id})
                         except Exception as save_e: print(f"Error saving unidentified face {unique_id}: {save_e}")


                # --- Save Final Attendance (Unique Identified Students) ---
                attendance_results = []
                absent_students_list = []
                with transaction.atomic():
                    for student in students_in_course:
                        status = "Present" if student.id in present_student_ids else "Absent"
                        if status == "Absent": absent_students_list.append({'id': student.id, 'name': student.name, 'roll_no': student.roll_no})
                        attendance_results.append({'name': student.name, 'roll_no': student.roll_no, 'status': status})
                        try: Attendance.objects.update_or_create(date=attendance_date, course=selected_course, student=student, defaults={'status': status})
                        except Exception as att_e: print(f"ERROR saving final attendance: {att_e}")

                # --- Set messages and context ---
                message = f"Processed {processed_files_count} photo(s). Initial attendance saved. Identified {len(present_student_ids)} unique student(s)."
                processed_data = {"results": attendance_results, "date": attendance_date_str, "course_name": selected_course.name}

                if unidentified_face_details:
                    manual_correction_data = { "unidentified_faces": unidentified_face_details, "absent_students": absent_students_list, "correction_date": attendance_date_str, "correction_course_id": selected_course.id }
                    messages.warning(request, f"{len(unidentified_face_details)} detected face images require manual review.")
                elif len(all_extracted_faces_info) > 0 :
                     messages.info(request, "All detected faces were identified.")
                elif len(all_extracted_faces_info) == 0:
                     messages.warning(request, "No faces were detected in the uploaded photo(s) by DeepFace.")


        # --- Outer Error Handling ---
        except Course.DoesNotExist: error = "Selected course not found."
        except FileNotFoundError as fnf_e: error = f"File not found error: {fnf_e}."
        except ValueError as ve: error = f"Invalid data format provided (e.g., date): {ve}"
        except Exception as e: error = f"An unexpected error occurred: {e}"; print(f"[ERROR] {e}")

        # --- Add messages and prepare final context for rendering ---
        if error: messages.error(request, error)
        if message and not error: messages.success(request, message)

        context = {
            'courses': courses, 'today': date.today().strftime('%Y-%m-%d'),
            'processed_data': processed_data, 'manual_correction_data': manual_correction_data,
            'processed_course_id': processed_course_id
        }
        return render(request, 'instructor_dashboard.html', context)

    # --- Handle GET Request ---
    else:
        return render(request, 'instructor_dashboard.html', context)

def load_subjects(request):
    course_id = request.GET.get('course_id')
    subjects_data = []
    if course_id:
        try:
            subjects = Subject.objects.filter(course_id=course_id).order_by('name')
            subjects_data = list(subjects.values('id', 'name'))
        except Exception as e:
            print(f"Error loading subjects: {e}")
            return JsonResponse({'error': 'Could not load subjects'}, status=500)
    return JsonResponse({'subjects': subjects_data})

def manage_courses_view(request):
    courses = Course.objects.prefetch_related('students').all()
    context = {'courses': courses}
    return render(request, 'manage_courses.html', context)


def manage_course_students_view(request, course_id):
    course = get_object_or_404(Course, pk=course_id)
    from .forms import StudentForm

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'remove':
            student_id_to_remove = request.POST.get('student_id')
            if not student_id_to_remove:
                 messages.error(request, "Invalid request: Missing student ID to remove.")
            else:
                try:
                    student_to_remove = Students.objects.get(id=student_id_to_remove)
                    course.students.remove(student_to_remove)
                    # --- Optional: Remove photo from student_photo_db ---
                    db_photo_path_base = os.path.join(settings.MEDIA_ROOT, 'student_photo_db', str(course_id), str(student_to_remove.id))
                    for ext in ['.jpg', '.jpeg', '.png']:
                        db_photo_path = db_photo_path_base + ext
                        if os.path.exists(db_photo_path):
                            try:
                                os.remove(db_photo_path)
                                print(f"Removed photo from DeepFace DB: {db_photo_path}")
                                break
                            except OSError as e: print(f"Error removing photo from DeepFace DB: {e}")
                    # --- End Optional ---
                    messages.success(request, f"Successfully removed '{student_to_remove.name}' from course '{course.name}'.")
                except Students.DoesNotExist:
                     messages.error(request, "Student to remove not found.")
                except Exception as e:
                     messages.error(request, f"Error removing student: {e}")
                     print(f"Error removing student {student_id_to_remove} from course {course_id}: {e}")
            return redirect('manage_course_students', course_id=course.id)

        elif action == 'add_new':
            student_name = request.POST.get('student_name', '').strip()
            roll_no = request.POST.get('roll_no', '').strip()
            student_image_file = request.FILES.get('student_image')

            if not student_name or not roll_no:
                messages.error(request, "Student Name and Roll Number are required.")
            else:
                photo_relative_path_for_model = None
                try:
                    student, created = Students.objects.get_or_create(
                        roll_no=roll_no,
                        defaults={'name': student_name}
                    )
                    if not created and student.name != student_name:
                        student.name = student_name
                        student.save(update_fields=['name'])

                    if student_image_file:
                        target_db_dir = os.path.join(settings.MEDIA_ROOT, 'student_photo_db', str(course_id))
                        os.makedirs(target_db_dir, exist_ok=True)
                        original_filename = student_image_file.name
                        extension = os.path.splitext(original_filename)[1].lower()
                        if not extension in ['.jpg', '.jpeg', '.png']:
                             raise ValueError("Invalid file type. Please upload JPG, JPEG, or PNG images.")

                        new_filename = f"{student.id}{extension}" # Use STUDENT ID
                        full_destination_path = os.path.join(target_db_dir, new_filename)

                        with open(full_destination_path, 'wb+') as destination:
                             for chunk in student_image_file.chunks():
                                 destination.write(chunk)

                        photo_relative_path_for_model = os.path.join('student_photo_db', str(course_id), new_filename).replace(os.sep, "/")
                        print(f"Saved/Updated photo in DeepFace DB: {full_destination_path}")
                        print(f"Relative path for Model: {photo_relative_path_for_model}")

                        if student.photo != photo_relative_path_for_model:
                           student.photo = photo_relative_path_for_model
                           student.save(update_fields=['photo'])
                           photo_saved_msg = "Photo updated."
                        else:
                            photo_saved_msg = "Photo processed (already up-to-date)."
                    else:
                         photo_saved_msg = "No new photo provided."

                    action_msg = "Created new student" if created else "Updated existing student"
                    messages.info(request, f"{action_msg} '{student.name}' ({roll_no}). {photo_saved_msg}")

                    if not course.students.filter(pk=student.id).exists():
                        course.students.add(student)
                        messages.success(request, f"Successfully linked student '{student.name}' to course '{course.name}'.")
                    else:
                        messages.warning(request, f"Student '{student.name}' is already linked to this course.")

                except ValueError as ve:
                      messages.error(request, str(ve))
                except IntegrityError as e:
                     messages.error(request, f"Database integrity error for roll no {roll_no}: {e}")
                except Exception as e:
                    print(f"Error adding new student: {e}")
                    messages.error(request, f"An error occurred while adding the student: {e}")

            return redirect('manage_course_students', course_id=course.id)

        else:
            messages.error(request, "Invalid action specified.")
            return redirect('manage_course_students', course_id=course.id)

    else:
        linked_students = course.students.all().order_by('name')
        context = {
            'course': course,
            'linked_students': linked_students,
        }
        return render(request, 'manage_students.html', context)


@require_POST
def delete_course_view(request, course_id):
    course = get_object_or_404(Course, id=course_id)
    try:
        course_name = course.name
        course_db_dir = os.path.join(settings.MEDIA_ROOT, 'student_photo_db', str(course_id))
        if os.path.exists(course_db_dir):
           try:
               shutil.rmtree(course_db_dir)
               print(f"Removed DeepFace DB folder: {course_db_dir}")
           except OSError as e:
               print(f"Error removing DeepFace DB folder {course_db_dir}: {e}")

        course.delete()
        messages.success(request, f'Successfully deleted course "{course_name}".')
    except Exception as e:
        messages.error(request, f"Error deleting course '{course.name}': {e}")
        print(f"Error deleting course ID {course_id}: {e}")
    return redirect('manage_courses')

def export_attendance_excel(request, course_id):
    try:
        course = get_object_or_404(Course, id=course_id)
        students = course.students.all().order_by('roll_no')
        if not students.exists():
             messages.error(request, f"No students found linked to course '{course.name}'.")
             return redirect('instructor_dashboard')

        dates = Attendance.objects.filter(course=course).values_list('date', flat=True).distinct().order_by('date')

        if not dates:
             messages.error(request, f"No attendance records found for course '{course.name}'.")
             return redirect('instructor_dashboard')

        wb = Workbook()
        ws = wb.active
        ws.title = f"{course.name} Attendance"

        headers = ["Roll No", "Student Name"] + [d.strftime('%Y-%m-%d') for d in dates]
        for col_num, header_title in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            cell = ws[f"{col_letter}1"]
            cell.value = header_title
            cell.font = cell.font.copy(bold=True)

        attendance_records = Attendance.objects.filter(
            course=course,
            student__in=students,
            date__in=dates
        ).select_related('student')

        attendance_dict = {}
        for record in attendance_records:
            if record.student_id not in attendance_dict:
                attendance_dict[record.student_id] = {}
            attendance_dict[record.student_id][record.date] = record.status

        for row_num, student in enumerate(students, 2):
            ws[f"A{row_num}"] = student.roll_no
            ws[f"B{row_num}"] = student.name
            student_attendance = attendance_dict.get(student.id, {})
            for col_num, attendance_date in enumerate(dates, 3):
                status = student_attendance.get(attendance_date, "N/A")
                col_letter = get_column_letter(col_num)
                ws[f"{col_letter}{row_num}"] = status

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        filename = f"attendance_{course.name.replace(' ', '_')}_{date.today().strftime('%Y%m%d')}.xlsx"
        safe_filename = "".join(c for c in filename if c.isalnum() or c in ('_', '-')).rstrip()
        response['Content-Disposition'] = f'attachment; filename="{safe_filename}"'

        wb.save(response)
        return response

    except Course.DoesNotExist:
        messages.error(request, "Course not found.")
        return redirect('instructor_dashboard')
    except Exception as e:
        messages.error(request, f"An error occurred during export: {e}")
        print(f"Error exporting attendance for course {course_id}: {e}")
        return redirect('instructor_dashboard')